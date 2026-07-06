"""
前瞻数据模块：读取 looking_ahead.xlsx，输出结构化的比赛前瞻分析。

looking_ahead.xlsx 结构：
- 每场比赛为一个数据块，以"国家"行开头
- 后续行为维度分析（实力、身价、状态、伤病等），维度名非固定
- 数据块之间以空行分隔
"""

import os
import json
from typing import Dict, List, Any


def _try_read_xlsx(filepath: str) -> List[List[str]]:
    """读取 xlsx 全部行，返回 [[col1, col2, col3], ...]"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(filepath)
    ws = wb["Sheet1"]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        rows.append([str(c).strip() if c is not None else "" for c in row])
    return rows


def _parse_blocks(rows: List[List[str]]) -> List[Dict[str, Any]]:
    """
    将行列表解析为比赛数据块。
    每个块以"国家"行开头，包含团队名称和若干分析维度。
    """
    blocks = []
    current_block = None

    for row in rows:
        # 跳过全空行
        if all(c == "" for c in row):
            if current_block is not None:
                blocks.append(current_block)
                current_block = None
            continue

        label = row[0] if len(row) > 0 else ""
        col1 = row[1] if len(row) > 1 else ""
        col2 = row[2] if len(row) > 2 else ""

        if label == "国家":
            # 新块开始
            if current_block is not None:
                blocks.append(current_block)
            current_block = {
                "teamA": col1,
                "teamB": col2,
                "analysis": {},
            }
        elif current_block is not None and label:
            current_block["analysis"][label] = {
                "teamA": col1,
                "teamB": col2,
            }

    # 最后一个块
    if current_block is not None:
        blocks.append(current_block)

    return blocks


def _parse_legacy_txt(text: str) -> List[Dict[str, Any]]:
    """
    兼容旧的纯文本格式（如果 xlsx 不可用）。
    格式：每 8 行一个比赛块（抬头 + 7 个分析维度）。
    """
    lines = [line.strip() for line in text.split("\n") if line.strip()]
    block_size = 8
    blocks = []

    for i in range(0, len(lines), block_size):
        block_lines = lines[i : i + block_size]
        if len(block_lines) < 8:
            continue

        # 第一行是 "国家, 球队A, 球队B" 格式
        header = block_lines[0].split(",")
        if len(header) < 3:
            continue
        team_a = header[1].strip()
        team_b = header[2].strip()

        analysis = {}
        for line in block_lines[1:]:
            parts = line.split(",")
            if len(parts) >= 3:
                dim = parts[0].strip()
                analysis[dim] = {
                    "teamA": parts[1].strip(),
                    "teamB": parts[2].strip(),
                }

        blocks.append({
            "teamA": team_a,
            "teamB": team_b,
            "analysis": analysis,
        })

    return blocks


def load_match_analysis(data_dir: str) -> List[Dict[str, Any]]:
    """
    加载比赛前瞻分析数据。优先读取 xlsx，回退到 txt。
    返回：[{teamA, teamB, analysis: {dimension: {teamA, teamB}}}, ...]
    """
    xlsx_path = os.path.join(data_dir, "looking_ahead.xlsx")

    if os.path.exists(xlsx_path):
        try:
            rows = _try_read_xlsx(xlsx_path)
            blocks = _parse_blocks(rows)
            if blocks:
                return blocks
        except Exception as e:
            print(f"  [警告] xlsx 读取失败 ({e})，尝试 txt 回退")

    # 回退 txt
    txt_path = os.path.join(data_dir, "looking_ahead.txt")
    if os.path.exists(txt_path):
        with open(txt_path, "r", encoding="utf-8") as f:
            return _parse_legacy_txt(f.read())

    print("  [警告] 未找到 looking_ahead.xlsx 或 .txt，前瞻数据为空")
    return []


def get_match_analysis_for_match(
    match_analysis: List[Dict[str, Any]],
    team_a: str,
    team_b: str,
) -> Dict[str, Any] | None:
    """
    根据两队名称查找对应的前瞻分析。
    支持双向匹配（A vs B 和 B vs A）。
    """
    for block in match_analysis:
        if (block["teamA"] == team_a and block["teamB"] == team_b) or \
           (block["teamA"] == team_b and block["teamB"] == team_a):
            return block
    return None


def load_and_export(data_dir: str, output_path: str) -> List[Dict[str, Any]]:
    """
    读取 xlsx 并导出为 JSON（供前端直接消费）。
    """
    blocks = load_match_analysis(data_dir)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(blocks, f, ensure_ascii=False, indent=2)
    print(f"  - match_analysis.json -> {output_path}")
    return blocks
